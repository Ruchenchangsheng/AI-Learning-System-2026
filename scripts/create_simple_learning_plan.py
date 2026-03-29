#!/usr/bin/env python3
"""
创建简单的学习计划Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_learning_plan_excel():
    """创建学习计划Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "学习计划总览"
    
    # 设置标题
    title_font = Font(name='微软雅黑', size=18, bold=True, color='000000')
    ws['A1'] = "AI学习计划 (2026.3.29 - 2026.6.30)"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 基本信息
    ws['A3'] = "学生: 爸爸"
    ws['A4'] = "AI助手: 小爪"
    ws['A5'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # 学习阶段表
    ws['A7'] = "📅 学习阶段计划"
    ws['A7'].font = Font(name='微软雅黑', size=14, bold=True, color='366092')
    
    headers = ["阶段", "时间范围", "持续时间", "主要目标", "关键成果", "状态"]
    data = [
        ["第1阶段: 基础速成", "3.29-4.30", "33天", "Python基础 + 机器学习入门", "掌握Python编程，理解ML基础", "进行中"],
        ["第2阶段: 深度学习专精", "5.1-5.31", "31天", "深度学习原理 + 计算机视觉", "掌握深度学习，完成CV项目", "未开始"],
        ["第3阶段: 项目冲刺", "6.1-6.20", "20天", "全栈AI应用开发", "完成完整AI项目", "未开始"],
        ["第4阶段: 求职准备", "6.21-6.30", "10天", "面试训练 + 简历优化", "拿到AI工程师Offer", "未开始"]
    ]
    
    # 写入表头
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据
    for row_idx, row_data in enumerate(data, 10):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # 调整列宽
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        for row in range(9, 9 + len(data) + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(9, 9 + len(data) + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border
    
    # 每日安排表
    ws2 = wb.create_sheet(title="每日安排")
    
    ws2['A1'] = "⏰ 每日学习安排"
    ws2['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='366092')
    
    schedule_headers = ["时间段", "活动", "时长", "地点", "工具"]
    schedule_data = [
        ["08:00-09:00", "查看今日计划，准备学习", "1小时", "书房", "电脑、笔记本"],
        ["09:00-12:00", "上午学习（Python/ML）", "3小时", "书房", "VS Code、教程"],
        ["12:00-14:00", "午休", "2小时", "餐厅", ""],
        ["14:00-18:00", "下午学习（练习/项目）", "4小时", "书房", "Python、练习题目"],
        ["18:00-20:00", "晚餐休息", "2小时", "家里", ""],
        ["20:00-21:00", "晚上学习（项目实践）", "1小时", "书房", "项目代码"],
        ["21:00-21:30", "模拟面试（小爪提问）", "30分钟", "书房", "QQ、笔记"],
        ["22:00-22:30", "总结提交（GitHub）", "30分钟", "书房", "Git、GitHub"]
    ]
    
    # 写入每日安排
    for col_idx, header in enumerate(schedule_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_idx, row_data in enumerate(schedule_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # 调整列宽
    for col_idx in range(1, len(schedule_headers) + 1):
        max_length = 0
        for row in range(3, 3 + len(schedule_data) + 1):
            cell_value = ws2.cell(row=row, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 30)
        ws2.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # 添加边框
    for row in range(3, 3 + len(schedule_data) + 1):
        for col in range(1, len(schedule_headers) + 1):
            ws2.cell(row=row, column=col).border = thin_border
    
    # 里程碑表
    ws3 = wb.create_sheet(title="里程碑")
    
    ws3['A1'] = "🎯 重要里程碑"
    ws3['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='366092')
    
    milestone_headers = ["里程碑", "目标日期", "完成标准", "重要性", "状态"]
    milestone_data = [
        ["完成Python基础", "2026-04-05", "掌握变量、数据类型、控制流、函数", "高", "进行中"],
        ["完成第1个小项目", "2026-04-03", "完成简易计算器项目", "高", "待开始"],
        ["掌握机器学习基础", "2026-04-30", "理解常见ML算法和原理", "中", "未开始"],
        ["完成深度学习入门", "2026-05-15", "掌握神经网络基本原理", "中", "未开始"],
        ["完成CV项目", "2026-05-31", "完成计算机视觉项目", "高", "未开始"],
        ["完成完整AI项目", "2026-06-20", "完成全栈AI应用", "高", "未开始"],
        ["通过模拟面试", "2026-06-25", "面试评分达到80分以上", "中", "未开始"],
        ["拿到Offer", "2026-06-30", "获得AI工程师职位", "最高", "未开始"]
    ]
    
    # 写入里程碑
    for col_idx, header in enumerate(milestone_headers, 1):
        cell = ws3.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_idx, row_data in enumerate(milestone_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # 调整列宽
    for col_idx in range(1, len(milestone_headers) + 1):
        max_length = 0
        for row in range(3, 3 + len(milestone_data) + 1):
            cell_value = ws3.cell(row=row, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 40)
        ws3.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # 添加边框
    for row in range(3, 3 + len(milestone_data) + 1):
        for col in range(1, len(milestone_headers) + 1):
            ws3.cell(row=row, column=col).border = thin_border
    
    # 保存文件
    output_file = "excel_files/学习计划.xlsx"
    wb.save(output_file)
    print(f"✅ 学习计划Excel已生成: {output_file}")
    
    return True

if __name__ == "__main__":
    create_learning_plan_excel()