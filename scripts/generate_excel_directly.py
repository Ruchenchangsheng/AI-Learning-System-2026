#!/usr/bin/env python3
"""
直接生成Excel文件（不依赖pandas）
使用openpyxl基础功能创建Excel
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

class DirectExcelGenerator:
    def __init__(self, base_path="."):
        self.base_path = base_path
        self.config_path = os.path.join(base_path, "config")
        self.progress_path = os.path.join(base_path, "progress")
        self.excel_path = os.path.join(base_path, "excel_files")
        
        # 创建Excel目录
        os.makedirs(self.excel_path, exist_ok=True)
    
    def create_styled_workbook(self, title):
        """创建带样式的Workbook"""
        wb = Workbook()
        ws = wb.active
        ws.title = "封面"
        
        # 设置标题样式
        title_font = Font(name='微软雅黑', size=20, bold=True, color='000000')
        ws['A1'] = title
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 合并单元格
        ws.merge_cells('A1:E1')
        
        # 添加生成时间
        time_font = Font(name='微软雅黑', size=10, color='666666')
        ws['A3'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = time_font
        
        return wb
    
    def add_data_sheet(self, wb, sheet_name, data, headers):
        """添加数据工作表"""
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
        
        # 设置表头样式
        header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # 调整列宽
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        return ws
    
    def generate_learning_plan_excel(self):
        """生成学习计划Excel"""
        json_file = os.path.join(self.config_path, "learning_plan.json")
        excel_file = os.path.join(self.excel_path, "学习计划.xlsx")
        
        if not os.path.exists(json_file):
            print(f"❌ JSON文件不存在: {json_file}")
            return False
        
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 创建Workbook
            wb = self.create_styled_workbook("AI学习计划总览")
            
            # 1. 总体目标
            if 'overall_goal' in data:
                goal = data['overall_goal']
                goal_data = [
                    ["项目", "内容"],
                    ["目标名称", goal.get('name', '')],
                    ["时间范围", goal.get('time_range', '')],
                    ["总体目标", goal.get('description', '')],
                    ["成功标准", goal.get('success_criteria', '')],
                    ["预期成果", goal.get('expected_outcomes', '')]
                ]
                self.add_data_sheet(wb, "总体目标", goal_data, ["项目", "内容"])
            
            # 2. 学习阶段
            if 'learning_phases' in data:
                phases_data = [["阶段名称", "时间范围", "持续时间", "主要目标", "关键成果", "优先级"]]
                for phase in data['learning_phases']:
                    phases_data.append([
                        phase.get('name', ''),
                        phase.get('time_range', ''),
                        phase.get('duration', ''),
                        phase.get('main_goals', ''),
                        phase.get('key_outcomes', ''),
                        phase.get('priority', '')
                    ])
                self.add_data_sheet(wb, "学习阶段", phases_data, phases_data[0])
            
            # 3. 每日安排
            if 'daily_schedule' in data:
                schedule_data = [["时间段", "活动", "时长", "地点", "工具"]]
                for slot in data['daily_schedule']:
                    schedule_data.append([
                        slot.get('time', ''),
                        slot.get('activity', ''),
                        slot.get('duration', ''),
                        slot.get('location', ''),
                        slot.get('tools', '')
                    ])
                self.add_data_sheet(wb, "每日安排", schedule_data, schedule_data[0])
            
            # 4. 里程碑
            if 'milestones' in data:
                milestones_data = [["里程碑", "目标日期", "完成标准", "重要性", "状态"]]
                for milestone in data['milestones']:
                    milestones_data.append([
                        milestone.get('name', ''),
                        milestone.get('target_date', ''),
                        milestone.get('completion_criteria', ''),
                        milestone.get('importance', ''),
                        milestone.get('status', '未开始')
                    ])
                self.add_data_sheet(wb, "里程碑", milestones_data, milestones_data[0])
            
            # 5. 学习资源
            if 'resources' in data:
                resources_data = [["资源类型", "名称", "链接", "描述", "推荐程度"]]
                for resource in data['resources']:
                    resources_data.append([
                        resource.get('type', ''),
                        resource.get('name', ''),
                        resource.get('url', ''),
                        resource.get('description', ''),
                        resource.get('recommendation_level', '')
                    ])
                self.add_data_sheet(wb, "学习资源", resources_data, resources_data[0])
            
            # 保存Excel文件
            wb.save(excel_file)
            print(f"✅ 学习计划Excel已生成: {excel_file}")
            return True
            
        except Exception as e:
            print(f"❌ 生成失败: {e}")
            return False
    
    def generate_daily_checklist_excel(self):
        """生成每日任务模板Excel"""
        json_file = os.path.join(self.config_path, "daily_checklist.json")
        excel_file = os.path.join(self.excel_path, "每日任务模板.xlsx")
        
        if not os.path.exists(json_file):
            print(f"❌ JSON文件不存在: {json_file}")
            return False
        
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 创建Workbook
            wb = self.create_styled_workbook("每日任务模板")
            
            # 1. 每日模板
            if 'daily_template' in data:
                template = data['daily_template']
                template_data = [
                    ["项目", "内容"],
                    ["日期", template.get('date', '')],
                    ["星期", template.get('day_of_week', '')],
                    ["阶段", template.get('phase', '')],
                    ["周数", str(template.get('week_number', 0))],
                    ["天数", str(template.get('day_number', 0))]
                ]
                
                # 上午任务
                morning = template.get('morning_session', {})
                template_data.append(["上午时间", morning.get('time', '')])
                template_data.append(["上午任务", '\n'.join(morning.get('tasks', []))])
                template_data.append(["上午完成", "是" if morning.get('completed', False) else "否"])
                template_data.append(["上午备注", morning.get('notes', '')])
                
                # 下午任务
                afternoon = template.get('afternoon_session', {})
                template_data.append(["下午时间", afternoon.get('time', '')])
                template_data.append(["下午任务", '\n'.join(afternoon.get('tasks', []))])
                template_data.append(["下午完成", "是" if afternoon.get('completed', False) else "否"])
                template_data.append(["下午备注", afternoon.get('notes', '')])
                
                # 晚上任务
                evening = template.get('evening_session', {})
                template_data.append(["晚上时间", evening.get('time', '')])
                template_data.append(["晚上任务", '\n'.join(evening.get('tasks', []))])
                template_data.append(["晚上完成", "是" if evening.get('completed', False) else "否"])
                template_data.append(["晚上备注", evening.get('notes', '')])
                
                self.add_data_sheet(wb, "每日模板", template_data, ["项目", "内容"])
            
            # 2. 第1阶段任务
            if 'phase1_tasks' in data:
                tasks_data = [["周", "天", "日期", "重点", "上午任务", "下午任务", "晚上任务"]]
                phase1 = data['phase1_tasks']
                
                for week_num, week_data in phase1.items():
                    for day_key, day_data in week_data.items():
                        tasks_data.append([
                            week_num,
                            day_key,
                            day_data.get('date', ''),
                            day_data.get('focus', ''),
                            '\n'.join(day_data.get('morning_tasks', [])),
                            '\n'.join(day_data.get('afternoon_tasks', [])),
                            '\n'.join(day_data.get('evening_tasks', []))
                        ])
                
                self.add_data_sheet(wb, "第1阶段任务", tasks_data, tasks_data[0])
            
            # 3. 检查清单
            if 'checklist_categories' in data:
                categories_data = [["类别", "序号", "检查项"]]
                categories = data['checklist_categories']
                
                for category_name, items in categories.items():
                    for i, item in enumerate(items, 1):
                        categories_data.append([category_name, str(i), item])
                
                self.add_data_sheet(wb, "检查清单", categories_data, categories_data[0])
            
            # 4. 完成指标
            if 'completion_metrics' in data:
                metrics_data = [["指标类型", "等级", "描述"]]
                metrics = data['completion_metrics']
                
                for metric_name, levels in metrics.items():
                    for level_name, description in levels.items():
                        metrics_data.append([metric_name, level_name, description])
                
                self.add_data_sheet(wb, "完成指标", metrics_data, metrics_data[0])
            
            # 5. 每日问题
            if 'daily_questions' in data:
                questions_data = [["时间段", "序号", "问题"]]
                questions = data['daily_questions']
                
                for time_period, q_list in questions.items():
                    for i, question in enumerate(q_list, 1):
                        questions_data.append([time_period, str(i), question])
                
                self.add_data_sheet(wb, "每日问题", questions_data, questions_data[0])
            
            # 保存Excel文件
            wb.save(excel_file)
            print(f"✅ 每日任务模板Excel已生成: {excel_file}")
            return True
            
        except Exception as e:
            print(f"❌ 生成失败: {e}")
            return False
    
    def generate_learning_stats_excel(self):
        """生成学习统计Excel"""
        json_file = os.path.join(self.progress_path, "learning_stats.json")
        excel_file = os.path.join(self.excel_path, "学习统计.xlsx")
        
        if not os.path.exists(json_file):
            print(f"❌ JSON文件不存在: {json_file}")
            return False
        
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 创建Workbook
            wb = self.create_styled_workbook("学习统计")
            
            # 1. 当前进度
            if 'current_progress' in data:
                progress_data = [["指标", "值"]]
                progress = data['current_progress']
                
                for key, value in progress.items():
                    progress_data.append([key, str(value)])
                
                self.add_data_sheet(wb, "当前进度", progress_data, progress_data[0])
            
            # 2. 学习统计
            if 'learning_statistics' in data:
                stats_data = [["统计项", "数值"]]
                stats = data['learning_statistics']
                
                for key, value in stats.items():
                    stats_data.append([key, str(value)])
                
                self.add_data_sheet(wb, "学习统计", stats_data, stats_data[0])
            
            # 3. 知识掌握
            if 'knowledge_mastery' in data:
                knowledge_data = [["学科", "掌握程度", "已学主题", "最后练习"]]
                knowledge = data['knowledge_mastery']
                
                for subject, info in knowledge.items():
                    knowledge_data.append([
                        subject,
                        str(info.get('level', 0)),
                        ', '.join(info.get('topics_covered', [])),
                        info.get('last_practiced', '')
                    ])
                
                self.add_data_sheet(wb, "知识掌握", knowledge_data, knowledge_data[0])
            
            # 4. 性能指标
            if 'performance_metrics' in data:
                performance_data = [["指标", "得分"]]
                metrics = data['performance_metrics']
                
                for key, value in metrics.items():
                    performance_data.append([key, str(value)])
                
                self.add_data_sheet(wb, "性能指标", performance_data, performance_data[0])
            
            # 5. 下一个里程碑
            if 'next_milestones' in data:
                milestones_data = [["里程碑", "目标日期", "优先级", "状态"]]
                milestones = data['next_milestones']
                
                for milestone in milestones:
                    milestones_data.append([
                        milestone.get('milestone', ''),
                        milestone.get('target_date', ''),
                        milestone.get('priority', ''),
                        "待完成"
                    ])
                
                self.add_data_sheet(wb, "下一个里程碑", milestones_data, milestones_data[0])
            
            # 保存Excel文件
            wb.save(excel_file)
            print(f"✅ 学习统计Excel已生成: {excel_file}")
            return True
            
        except Exception as e:
            print(f"❌ 生成失败: {e}")
            return False
    
    def delete_json_files(self):
        """删除JSON文件"""
        json_files = [
            os.path.join(self.config_path, "learning_plan.json"),
            os.path.join(self.config_path, "daily_checklist.json"),
            os.path.join(self.progress_path, "learning_stats.json")
        ]
        
        deleted_count = 0
        for json_file in json_files:
            if os.path.exists(json_file):
                try:
                    os.remove(json_file)
                    print(f"🗑️ 已删除: {json_file}")
                    deleted_count += 1
                except Exception as e:
                    print(f"❌ 删除失败 {json_file}: {e}")
        
        return deleted_count
    
    def update_readme_for_excel(self):
        """更新README，说明使用Excel"""
        readme_file = os.path.join(self.base_path, "README.md")
        
        if not os.path.exists(readme_file):
            print(f"❌ README文件不存在: {readme_file}")
            return False
        
        try:
            with open(readme_file, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # 添加Excel说明部分
            excel_section = """

## 📊 Excel文件说明

系统已从JSON格式升级为Excel格式，提供更直观的数据查看方式：

### 生成的Excel文件：
1. **学习计划.xlsx** - 包含总体目标、学习阶段、每日安排、里程碑、学习资源
2. **每日任务模板.xlsx** - 包含每日模板、第1阶段任务、检查清单、完成指标、每日问题
3. **学习统计.xlsx** - 包含当前进度、学习统计、知识掌握、性能指标、下一个里程碑

### 使用方式：
- 直接使用Excel软件打开查看
- 数据更直观，支持筛选、排序、图表
- 无需安装额外依赖

### 文件位置：
所有Excel文件位于 `excel_files/` 目录中。

### 更新说明：
- ✅ 已删除原有的JSON配置文件
- ✅ 所有数据已转换为Excel格式
- ✅ 保持相同的结构和内容
- ✅ 更友好的查看和编辑体验
"""
            
            # 在文件末尾添加Excel说明
            if "## 📊 Excel文件说明" not in content:
                content += excel_section
            
            with open(readme_file, 'w', encoding='utf-8') as f:
                f.write(content)
            
            print("✅ README已更新，添加Excel说明")
            return True
            
        except Exception as e:
            print(f"❌ 更新README失败: {e}")
            return False
    
    def run_all(self):
        """执行所有步骤"""
        print("=" * 50)
        print("🚀 直接生成Excel文件（无需pandas依赖）")
        print("=" * 50)
        
        results = []
        
        # 1. 生成学习计划Excel
        print("\n1. 📋 生成学习计划Excel...")
        result1 = self.generate_learning_plan_excel()
        results.append(("学习计划", result1))
        
        # 2. 生成每日任务模板Excel
        print("\n2. 📋 生成每日任务模板Excel...")
        result2 = self.generate_daily_checklist_excel()
        results.append(("每日任务模板", result2))
        
        # 3. 生成学习统计Excel
        print("\n3. 📋 生成学习统计Excel...")
        result3 = self.generate_learning_stats_excel()
        results.append(("学习统计", result3))
        
        # 4. 删除JSON文件
        print("\n4. 🗑️ 删除JSON文件...")
        deleted_count = self.delete_json_files()
        results.append(("删除JSON", deleted_count > 0))
        
        # 5. 更新README
        print("\n5. 📝 更新README...")
        result5 = self.update_readme_for_excel()
        results.append(("更新README", result5))
        
        # 输出结果汇总
        print("\n" + "=" * 50)
        print("📈 执行结果汇总")
        print("=" * 50)
        
        success_count = 0
        for name, success in results:
            status = "✅ 成功" if success else "❌ 失败"
            print(f"{name}: {status}")
            if success:
                success_count += 1
        
        print(f"\n🎯 完成度: {success_count}/{len(results)}")
        
        if success_count > 0:
            print(f"\n📁 Excel文件保存在: {self.excel_path}")
            print("📋 生成的文件:")
            for file in os.listdir(self.excel_path):
                if file.endswith('.xlsx'):
                    print(f"  - {file}")
        
        return success_count == len(results)

def main():
    """主函数"""
    import sys
    
    # 检查是否在正确目录
    if not os.path.exists("config"):
        print("❌ 请在AI学习管理系统根目录运行此脚本")
        print("   或指定正确的路径: python generate_excel_directly.py /path/to/system")
        sys.exit(1)
    
    # 创建生成器实例
    generator = DirectExcelGenerator()
    
    # 执行生成
    success = generator.run_all()
    
    if success:
        print("\n🎉 所有Excel文件已直接生成！")
        print("📊 JSON文件已删除，系统已升级为Excel格式！")
        print("🚀 现在可以用Excel直接查看和管理学习数据了！")
    else:
        print("\n⚠️ 部分步骤失败，请检查错误信息")
    
    # 提供使用建议
    print("\n" + "=" * 50)
    print("💡 使用建议")
    print("=" * 50)
    print("1. 查看Excel: 打开 excel_files/ 目录中的文件")
    print("2. 无需安装: 不需要安装pandas等依赖")
    print("3. 直接使用: 用Excel软件打开即可")
    print("4. 后续更新: 小爪会直接更新Excel文件")
    
    return 0 if success else 1

if __name__ == "__main__":
    sys.exit(main())