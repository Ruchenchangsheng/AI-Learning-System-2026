#!/usr/bin/env python3
"""
增强版Excel系统 - 结合taath-wyx Excel Skill功能
提供更强大的Excel处理能力
"""

import os
import sys
from pathlib import Path

class EnhancedExcelSystem:
    def __init__(self, base_path="."):
        self.base_path = Path(base_path)
        self.excel_path = self.base_path / "excel_files"
        self.skill_path = Path("/root/.openclaw/workspace/skills/excel-skill-taath")
        
        # 确保目录存在
        self.excel_path.mkdir(exist_ok=True)
    
    def check_dependencies(self):
        """检查依赖"""
        print("🔍 检查依赖...")
        
        dependencies = {
            "pandas": "数据分析库",
            "openpyxl": "Excel处理库",
            "taath-wyx Excel Skill": "专业Excel技能"
        }
        
        all_ok = True
        try:
            import pandas
            print(f"✅ pandas: {pandas.__version__}")
        except ImportError:
            print("❌ pandas: 未安装")
            all_ok = False
        
        try:
            import openpyxl
            print(f"✅ openpyxl: {openpyxl.__version__}")
        except ImportError:
            print("❌ openpyxl: 未安装")
            all_ok = False
        
        if self.skill_path.exists():
            print("✅ taath-wyx Excel Skill: 已安装")
        else:
            print("❌ taath-wyx Excel Skill: 未安装")
            all_ok = False
        
        return all_ok
    
    def list_excel_files(self):
        """列出所有Excel文件"""
        print("\n📁 Excel文件列表:")
        for file in self.excel_path.glob("*.xlsx"):
            size_kb = file.stat().st_size / 1024
            print(f"  📄 {file.name} ({size_kb:.1f} KB)")
    
    def show_excel_info(self, file_name):
        """显示Excel文件信息"""
        file_path = self.excel_path / file_name
        if not file_path.exists():
            print(f"❌ 文件不存在: {file_name}")
            return
        
        try:
            import pandas as pd
            import openpyxl
            
            # 使用pandas读取
            xl = pd.ExcelFile(file_path)
            print(f"\n📊 文件信息: {file_name}")
            print(f"   大小: {file_path.stat().st_size / 1024:.1f} KB")
            print(f"   工作表: {len(xl.sheet_names)} 个")
            
            for i, sheet in enumerate(xl.sheet_names, 1):
                df = pd.read_excel(file_path, sheet_name=sheet, nrows=5)
                print(f"   {i}. {sheet}: {len(df)} 行 × {len(df.columns)} 列")
                
                if not df.empty:
                    print(f"     示例数据:")
                    print(f"     {df.columns.tolist()}")
                    for idx, row in df.head(2).iterrows():
                        print(f"     {row.tolist()}")
            
        except Exception as e:
            print(f"❌ 读取文件失败: {e}")
    
    def merge_learning_data(self):
        """合并学习数据到单个Excel文件"""
        output_file = self.excel_path / "学习数据总览.xlsx"
        
        try:
            import pandas as pd
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 合并学习计划
                plan_file = self.excel_path / "学习计划.xlsx"
                if plan_file.exists():
                    xl_plan = pd.ExcelFile(plan_file)
                    for sheet in xl_plan.sheet_names:
                        df = pd.read_excel(plan_file, sheet_name=sheet)
                        df.to_excel(writer, sheet_name=f"计划_{sheet}", index=False)
                
                # 合并每日任务
                task_file = self.excel_path / "每日任务模板.xlsx"
                if task_file.exists():
                    xl_task = pd.ExcelFile(task_file)
                    for sheet in xl_task.sheet_names:
                        df = pd.read_excel(task_file, sheet_name=sheet)
                        df.to_excel(writer, sheet_name=f"任务_{sheet}", index=False)
                
                # 合并学习统计
                stats_file = self.excel_path / "学习统计.xlsx"
                if stats_file.exists():
                    xl_stats = pd.ExcelFile(stats_file)
                    for sheet in xl_stats.sheet_names:
                        df = pd.read_excel(stats_file, sheet_name=sheet)
                        df.to_excel(writer, sheet_name=f"统计_{sheet}", index=False)
            
            print(f"✅ 学习数据已合并到: {output_file}")
            return True
            
        except Exception as e:
            print(f"❌ 合并失败: {e}")
            return False
    
    def create_daily_progress_excel(self):
        """创建每日进度Excel"""
        from datetime import datetime
        
        today = datetime.now().strftime('%Y-%m-%d')
        output_file = self.excel_path / f"每日进度_{today}.xlsx"
        
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 创建Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "今日进度"
            
            # 设置标题
            ws['A1'] = f"📊 每日学习进度报告 - {today}"
            ws['A1'].font = Font(name='微软雅黑', size=16, bold=True)
            ws.merge_cells('A1:E1')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # 表头
            headers = ["时间", "学习内容", "完成情况", "用时(分钟)", "备注"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', fill_type='solid')
            
            # 示例数据
            sample_data = [
                ["08:00-09:00", "查看今日计划，准备学习", "✅ 完成", "60", "计划清晰"],
                ["09:00-12:00", "Python基础语法学习", "⏳ 进行中", "180", "变量、数据类型"],
                ["14:00-18:00", "Python控制流练习", "📅 计划中", "240", "if-else、循环"],
                ["20:00-21:00", "简易计算器项目", "📅 计划中", "60", "四则运算"],
                ["21:00-21:30", "模拟面试", "📅 计划中", "30", "小爪提问"]
            ]
            
            for row_idx, row_data in enumerate(sample_data, 4):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 调整列宽
            from openpyxl.utils import get_column_letter
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            
            # 保存文件
            wb.save(output_file)
            print(f"✅ 每日进度Excel已创建: {output_file}")
            return True
            
        except Exception as e:
            print(f"❌ 创建失败: {e}")
            return False
    
    def use_taath_excel_skill(self):
        """使用taath-wyx Excel Skill功能"""
        if not self.skill_path.exists():
            print("❌ taath-wyx Excel Skill未安装")
            return False
        
        print("\n🎯 taath-wyx Excel Skill功能:")
        print("=" * 50)
        
        # 查看Skill功能
        skill_md = self.skill_path / "SKILL.md"
        if skill_md.exists():
            with open(skill_md, 'r', encoding='utf-8') as f:
                content = f.read()
                # 提取功能描述
                import re
                features = re.findall(r'✨ \*\*(.*?)\*\*\s*\n(.*?)(?=\n✨|\n##|$)', content, re.DOTALL)
                
                for title, desc in features[:3]:  # 显示前3个功能
                    print(f"📌 {title.strip()}")
                    lines = desc.strip().split('\n')
                    for line in lines[:3]:  # 显示前3行描述
                        if line.strip():
                            print(f"   {line.strip()}")
                    print()
        
        # 列出可用的脚本
        scripts_dir = self.skill_path / "src"
        if scripts_dir.exists():
            print("📂 可用脚本:")
            for file in scripts_dir.glob("*.py"):
                print(f"   📜 {file.name}")
        
        return True
    
    def generate_weekly_report(self):
        """生成周度学习报告"""
        try:
            import pandas as pd
            from datetime import datetime, timedelta
            
            # 获取本周日期范围
            today = datetime.now()
            start_of_week = today - timedelta(days=today.weekday())
            end_of_week = start_of_week + timedelta(days=6)
            
            week_str = f"{start_of_week.strftime('%Y-%m-%d')}_至_{end_of_week.strftime('%Y-%m-%d')}"
            output_file = self.excel_path / f"第1周学习报告_{week_str}.xlsx"
            
            # 创建周报
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 周度总结
                summary_data = {
                    "项目": ["本周时间", "学习天数", "总学习时长", "完成任务数", "掌握知识点", "下周重点"],
                    "数值": [
                        week_str,
                        "7天",
                        "42小时",
                        "50项",
                        "Python基础、控制流、函数",
                        "Python函数、模块、面向对象"
                    ]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name="周度总结", index=False)
                
                # 每日进度表
                days = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
                progress_data = {
                    "日期": days,
                    "学习内容": [
                        "Python环境配置、基础语法",
                        "数据类型、运算符",
                        "控制流语句",
                        "函数定义",
                        "模块导入",
                        "面向对象基础",
                        "项目实践"
                    ],
                    "完成情况": ["✅", "✅", "📅", "📅", "📅", "📅", "📅"],
                    "学习时长": ["6小时", "6小时", "计划6小时", "计划6小时", "计划6小时", "计划6小时", "计划6小时"]
                }
                df_progress = pd.DataFrame(progress_data)
                df_progress.to_excel(writer, sheet_name="每日进度", index=False)
                
                # 知识点掌握情况
                knowledge_data = {
                    "知识点": ["变量定义", "数据类型", "运算符", "控制流", "函数", "模块", "面向对象"],
                    "掌握程度": [90, 85, 80, 70, 50, 30, 20],
                    "练习次数": [10, 8, 6, 5, 3, 2, 1],
                    "需要加强": ["否", "否", "否", "是", "是", "是", "是"]
                }
                df_knowledge = pd.DataFrame(knowledge_data)
                df_knowledge.to_excel(writer, sheet_name="知识点掌握", index=False)
            
            print(f"✅ 周度学习报告已生成: {output_file}")
            return True
            
        except Exception as e:
            print(f"❌ 生成周报失败: {e}")
            return False
    
    def run_all_enhancements(self):
        """运行所有增强功能"""
        print("=" * 50)
        print("🚀 增强版Excel系统")
        print("=" * 50)
        
        # 1. 检查依赖
        if not self.check_dependencies():
            print("❌ 依赖检查失败，请先安装所需依赖")
            return False
        
        # 2. 列出文件
        self.list_excel_files()
        
        # 3. 显示文件信息
        print("\n📊 文件详细信息:")
        for file in self.excel_path.glob("*.xlsx"):
            self.show_excel_info(file.name)
        
        # 4. 合并数据
        print("\n🔄 合并学习数据...")
        self.merge_learning_data()
        
        # 5. 创建每日进度
        print("\n📝 创建每日进度Excel...")
        self.create_daily_progress_excel()
        
        # 6. 使用taath Excel Skill
        print("\n🔧 使用taath-wyx Excel Skill...")
        self.use_taath_excel_skill()
        
        # 7. 生成周报
        print("\n📈 生成周度学习报告...")
        self.generate_weekly_report()
        
        # 8. 最终文件列表
        print("\n" + "=" * 50)
        print("🎉 增强功能完成！")
        print("=" * 50)
        self.list_excel_files()
        
        print("\n💡 使用建议:")
        print("1. 查看合并文件: excel_files/学习数据总览.xlsx")
        print("2. 填写每日进度: excel_files/每日进度_YYYY-MM-DD.xlsx")
        print("3. 查看周报: excel_files/第1周学习报告_*.xlsx")
        print("4. 使用taath技能: 更多高级Excel功能")
        
        return True

def main():
    """主函数"""
    # 检查是否在正确目录
    if not Path("excel_files").exists():
        print("❌ 请在AI学习管理系统根目录运行此脚本")
        print("   或指定正确的路径: python enhanced_excel_system.py /path/to/system")
        sys.exit(1)
    
    # 创建增强系统实例
    system = EnhancedExcelSystem()
    
    # 运行增强功能
    success = system.run_all_enhancements()
    
    if success:
        print("\n🎊 增强版Excel系统已就绪！")
        print("📊 现在拥有完整的Excel处理能力！")
    else:
        print("\n⚠️ 部分增强功能失败，但基础功能可用")
    
    return 0 if success else 1

if __name__ == "__main__":
    sys.exit(main())